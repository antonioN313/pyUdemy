from openpyxl.workbook import Workbook

wb = Workbook()
wb.create_sheet("First Sheet", 0)
wb.create_sheet("Middle Sheet", 2)
##sheet = wb.active()
##sheet.title = 'Planilha nro.1'
del wb['Sheet']
wb.save('pyExcel.xlsx')
