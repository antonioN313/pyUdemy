
import openpyxl
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world'
sheet.cell(row=1, column=2).value = 128
wb.save('pyExcel.xlsx')

sheet1 = wb.active

rows = {
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
}
for row in rows:
    sheet.append(row)
wb.save('pyAppend.xlsx')

sheet1 = wb.create_sheet('Sheet_Compras')
rows = (
    ('Compras', 'Valor (Kg) R$'),
    ('Banana', 8.90),
    ('Maçã Argentina', 2.45),
    ('Pão', 2.45),
    ('Iogurte', 9.90),
    ('Carne', 29.00),
    ('Feijão', 6.78),
    ('Arroz', 5.78),
)
for row in rows:
    sheet1.append(row)

cell_total = sheet1.cell(row=9, column=1)
cell_total.value = 'Total R$'
cell_sum = sheet1.cell(row=9, column=2)
cell_sum.value = '=SUM(B2:B8)'
wb.save('pyAppend.xlsx')
