from os import system, name
from time import sleep
import openpyxl


def clear():
    # for windows
    if name == 'nt':
        _ = system('cls')

    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system('clear')


def load_workbook(file_name):
    """Carrega o arquivo Excel."""
    return openpyxl.load_workbook(file_name)


def print_sheet_info(sheet):
    """Imprime informações sobre uma planilha."""
    print(sheet.title)
    print(sheet.max_row)
    print(sheet.max_column)


def print_cell_info(sheet):
    """Imprime informações sobre cada célula em uma planilha."""
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            print(f"Row {row_idx}, Column {col_idx}, Cell {cell.coordinate}: {cell.value}")


def print_range_info(sheet, start, end):
    """Imprime informações sobre um intervalo de células em uma planilha."""
    for rowOfCellObjects in sheet[start:end]:
        for cellObject in rowOfCellObjects:
            print(cellObject.coordinate, cellObject.value)
        print(f'\t END ROW {rowOfCellObjects} \t')


def main():
    """Função principal."""
    sleep(2)
    clear()
    wb = load_workbook('example.xlsx')
    print(type(wb))
    print(wb.sheetnames)
    sleep(4)
    clear()

    sheet1 = wb['Sheet1']
    sheet2 = wb['Sheet2']
    sheet3 = wb['Sheet3']

    print_sheet_info(sheet1)
    print_sheet_info(sheet2)
    print_sheet_info(sheet3)
    sleep(4)
    clear()

    anotherSheet = wb.active
    print(anotherSheet)
    sleep(4)

    print(sheet1.max_row, sheet2.max_row, sheet3.max_row)
    print(sheet1.max_column, sheet2.max_column, sheet3.max_column)
    sleep(4)
    clear()

    print_cell_info(sheet1)
    sleep(10)
    clear()

    print_range_info(sheet1, 'A1', 'C7')
    sleep(10)
    clear()


if __name__ == "__main__":
    main()
