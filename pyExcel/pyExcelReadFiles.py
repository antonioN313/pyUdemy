import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.sheetnames)


sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']
sheet3 = wb['Sheet3']
print(sheet1, sheet2, sheet3)
print(sheet1.title, sheet2.title, sheet3.title)

anotherSheet = wb.active
print(anotherSheet)

print(sheet1.max_row, sheet2.max_row, sheet3.max_row)
print(sheet1.max_column, sheet2.max_column, sheet3.max_column)

for row_idx, row in enumerate(sheet1.iter_rows(), start=1):
    for col_idx, cell in enumerate(row, start=1):
        print(f"Row {row_idx}, Column {col_idx}, Cell {cell.coordinate}: {cell.value}")


tuple_sheet1 = wb['Sheet1']

for rowOfCellObjects in tuple_sheet1['A1': 'C7']:
    for cellObject in rowOfCellObjects:
        print(cellObject.coordinate, cellObject.value)
    print(f'\t END ROW {rowOfCellObjects} \t')

